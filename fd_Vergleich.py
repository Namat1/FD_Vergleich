import io
from collections import Counter
from typing import Dict, List, Set, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Stammdaten
# ---------------------------------------------------------------------------

CUSTOMER_GROUPS: Dict[str, List[str]] = {
    "Malchow": [
        "115339", "215634", "216425", "216442", "216467", "216496", "216630", "216133",
        "216432", "216815", "216466", "216615", "219545", "219430", "216590", "215632",
        "216144", "216153", "219208", "216207", "216464", "216529", "216570", "216572",
        "216586", "216588", "216628", "216637", "216744", "219439", "216656", "215551",
        "219544", "216799", "216774", "216122", "216177", "216185", "216221", "216248",
        "216253", "216670", "216672", "219513", "216010", "216178", "216655", "216697",
        "216853", "216653", "216791", "216227", "216290", "216814", "216828", "219427",
        "219570", "216793", "216617", "215014", "215180", "216070", "219586", "216155",
        "216569", "216405", "216623", "219532", "219501", "210650", "216371",
    ],
    "Neumünster": [
        "213406", "214238", "214109", "210353", "211152", "217253", "210750", "210716",
        "214588", "214487", "218394", "210399", "214015", "210492", "218418", "211288",
        "211399", "213095", "218390", "211292", "218373", "218344", "213016", "210234",
        "210276", "218466", "218411", "218420", "218426", "218425", "218468", "218421",
        "214285", "214299", "214297", "214290", "218200", "218711", "218461", "210655",
        "210765", "218355", "210701", "213840", "218208", "211025",
    ],
    "Zarrentin": [
        "213568", "112681", "214289", "213458", "218601", "218804", "214321", "218801",
        "214094", "210509", "213580", "218707", "214376", "211380", "218867", "213553",
        "12823", "214296", "214043", "12923", "214192", "218607", "214590", "210455",
        "214001",
    ],
}

DAY_NAMES = {
    1: "Montag",
    2: "Dienstag",
    3: "Mittwoch",
    4: "Donnerstag",
    5: "Freitag",
    6: "Samstag",
}

DAY_COLUMNS_TOUR = {
    1: 6,
    2: 7,
    3: 8,
    4: 9,
    5: 10,
    6: 11,
}

SAP_COL_INDEX = 0
SAP_DAY_COL_INDEX = 6
TOUR_SAP_COL_INDEX = 1

LOCATION_ORDER = {name: index for index, name in enumerate(CUSTOMER_GROUPS.keys(), start=1)}
CUSTOMER_TO_LOCATION: Dict[str, str] = {}
CUSTOMER_TO_ORDER: Dict[str, int] = {}

for location_name, sap_list in CUSTOMER_GROUPS.items():
    for customer_index, sap_number in enumerate(sap_list, start=1):
        CUSTOMER_TO_LOCATION[sap_number] = location_name
        CUSTOMER_TO_ORDER[sap_number] = customer_index

SELECTED_SAPS: Set[str] = set(CUSTOMER_TO_LOCATION.keys())


# ---------------------------------------------------------------------------
# Helfer
# ---------------------------------------------------------------------------

def find_duplicate_saps() -> List[Tuple[str, List[str]]]:
    """Findet SAP-Nummern, die in mehreren Standorten oder mehrfach im selben
    Standort hinterlegt sind."""
    counter: Counter = Counter()
    location_map: Dict[str, List[str]] = {}
    for location, sap_list in CUSTOMER_GROUPS.items():
        for sap in sap_list:
            counter[sap] += 1
            location_map.setdefault(sap, []).append(location)
    return [(sap, location_map[sap]) for sap, count in counter.items() if count > 1]


def normalize_sap_series(series: pd.Series) -> pd.Series:
    """Normalisiert SAP-Nummern vektorisiert. Floats ohne Nachkomma werden zu int."""
    if series.empty:
        return series.astype(str)

    result = series.copy()
    # Numerische Werte -> int wenn möglich
    numeric = pd.to_numeric(result, errors="coerce")
    is_int = numeric.notna() & (numeric == numeric.round())

    out = result.astype(str)
    out = out.where(~is_int, numeric.where(is_int).astype("Int64").astype(str))
    out = out.str.strip()
    out = out.replace({"nan": "", "<NA>": "", "None": ""})
    return out


def read_sap_file(uploaded_file) -> Tuple[Dict[str, Set[int]], str, int]:
    """Liest die SAP-Datei: Spalte A = SAP, Spalte G = Liefertag 1..6."""
    excel = pd.ExcelFile(uploaded_file)
    sheet_name = excel.sheet_names[0]
    df = pd.read_excel(
        excel,
        sheet_name=sheet_name,
        header=0,
        usecols=[SAP_COL_INDEX, SAP_DAY_COL_INDEX],
        names=["sap", "tag"],
    )

    df["sap"] = normalize_sap_series(df["sap"])
    df["tag_num"] = pd.to_numeric(df["tag"], errors="coerce")

    mask = (
        df["sap"].ne("")
        & df["sap"].isin(SELECTED_SAPS)
        & df["tag_num"].between(1, 6, inclusive="both")
        & df["tag_num"].notna()
    )
    filtered = df.loc[mask, ["sap", "tag_num"]].copy()
    filtered["tag_int"] = filtered["tag_num"].astype(int)

    days_by_sap: Dict[str, Set[int]] = (
        filtered.groupby("sap")["tag_int"].agg(set).to_dict()
    )

    return days_by_sap, sheet_name, len(filtered)


def read_tourenplanung(uploaded_file) -> Tuple[pd.DataFrame, List[str]]:
    """Liest die ersten vier Blätter der Tourenplanung und gibt einen langen
    DataFrame zurück: eine Zeile pro (SAP, Tag, Blatt) mit gesetztem Wert."""
    excel = pd.ExcelFile(uploaded_file)
    sheet_names = excel.sheet_names[:4]

    usecols = [TOUR_SAP_COL_INDEX] + list(DAY_COLUMNS_TOUR.values())
    names = ["sap"] + [f"tag_{d}" for d in DAY_COLUMNS_TOUR.keys()]

    frames: List[pd.DataFrame] = []
    for sheet_name in sheet_names:
        df = pd.read_excel(
            excel,
            sheet_name=sheet_name,
            header=0,
            usecols=usecols,
            names=names,
        )
        if df.empty:
            continue
        df["sap"] = normalize_sap_series(df["sap"])
        df["blatt"] = sheet_name

        long = df.melt(
            id_vars=["sap", "blatt"],
            value_vars=[f"tag_{d}" for d in DAY_COLUMNS_TOUR.keys()],
            var_name="tag_col",
            value_name="wert",
        )
        long["tag_num"] = long["tag_col"].str.replace("tag_", "", regex=False).astype(int)
        long["wert_num"] = pd.to_numeric(long["wert"], errors="coerce")

        long = long[long["sap"].ne("") & long["wert_num"].notna()]
        frames.append(long[["sap", "blatt", "tag_num", "wert"]])

    if not frames:
        return pd.DataFrame(columns=["sap", "blatt", "tag_num", "wert"]), sheet_names

    return pd.concat(frames, ignore_index=True), sheet_names


def build_missing_in_sap(
    tour_df: pd.DataFrame,
    days_by_sap: Dict[str, Set[int]],
) -> pd.DataFrame:
    """Eine Zeile pro Kunde: welche Tage stehen in der Tourenplanung, fehlen aber
    in SAP als Liefertag."""
    if tour_df.empty:
        return _empty_result_df()

    known = tour_df[tour_df["sap"].isin(SELECTED_SAPS)].copy()
    if known.empty:
        return _empty_result_df()

    known["fehlt"] = known.apply(
        lambda row: row["tag_num"] not in days_by_sap.get(row["sap"], set()),
        axis=1,
    )
    missing = known[known["fehlt"]]
    if missing.empty:
        return _empty_result_df()

    agg = missing.groupby("sap", as_index=False).agg(
        tage=("tag_num", lambda x: sorted(set(x))),
        blaetter=("blatt", lambda x: sorted(set(x))),
    )

    agg["Standort"] = agg["sap"].map(CUSTOMER_TO_LOCATION).fillna("Ohne Zuordnung")
    agg["Fehlende Liefertage"] = agg["tage"].map(
        lambda tage: ", ".join(f"{d} {DAY_NAMES[d]}" for d in tage)
    )
    agg["Anzahl fehlender Tage"] = agg["tage"].map(len)
    agg["Blätter Tourenplanung"] = agg["blaetter"].map(", ".join)
    agg["Liefertage in SAP"] = agg["sap"].map(
        lambda s: ", ".join(f"{d} {DAY_NAMES[d]}" for d in sorted(days_by_sap.get(s, set())))
        or "(keine hinterlegt)"
    )
    agg["Hinweis"] = "Tage in Tourenplanung vorhanden, aber in SAP nicht als Liefertag hinterlegt"

    agg["_StandortSort"] = agg["Standort"].map(LOCATION_ORDER).fillna(999)
    agg["_KundenSort"] = agg["sap"].map(CUSTOMER_TO_ORDER).fillna(999999)
    agg = agg.rename(columns={"sap": "SAP Nummer"}).sort_values(
        ["_StandortSort", "_KundenSort"]
    ).reset_index(drop=True)

    return agg[_export_columns_missing()]


def build_missing_in_tour(
    tour_df: pd.DataFrame,
    days_by_sap: Dict[str, Set[int]],
) -> pd.DataFrame:
    """Eine Zeile pro Kunde: welche Tage sind in SAP als Liefertag hinterlegt,
    fehlen aber in der Tourenplanung."""
    days_in_tour: Dict[str, Set[int]] = {}
    if not tour_df.empty:
        days_in_tour = tour_df.groupby("sap")["tag_num"].agg(set).to_dict()

    rows: List[dict] = []
    for sap, sap_days in days_by_sap.items():
        tour_days = days_in_tour.get(sap, set())
        fehlend = sorted(sap_days - tour_days)
        if not fehlend:
            continue
        standort = CUSTOMER_TO_LOCATION.get(sap, "Ohne Zuordnung")
        rows.append({
            "Standort": standort,
            "SAP Nummer": sap,
            "Fehlende Liefertage": ", ".join(f"{d} {DAY_NAMES[d]}" for d in fehlend),
            "Anzahl fehlender Tage": len(fehlend),
            "Liefertage in Tourenplanung": ", ".join(
                f"{d} {DAY_NAMES[d]}" for d in sorted(tour_days)
            ) or "(nicht in Tourenplanung vorhanden)",
            "Hinweis": "Tage in SAP als Liefertag hinterlegt, kommen aber in der Tourenplanung nicht vor",
            "_StandortSort": LOCATION_ORDER.get(standort, 999),
            "_KundenSort": CUSTOMER_TO_ORDER.get(sap, 999999),
        })

    if not rows:
        return pd.DataFrame(columns=_export_columns_missing_tour())

    df = pd.DataFrame(rows)
    df = df.sort_values(["_StandortSort", "_KundenSort"]).reset_index(drop=True)
    return df[_export_columns_missing_tour()]


def build_unknown_saps(tour_df: pd.DataFrame) -> pd.DataFrame:
    """Eine Zeile pro unbekanntem Kunden: SAP-Nummern in der Tourenplanung,
    die in keiner Kundengruppe hinterlegt sind."""
    empty_cols = ["SAP Nummer", "Blätter Tourenplanung", "Gesamt-Vorkommen"]
    if tour_df.empty:
        return pd.DataFrame(columns=empty_cols)

    unknown = tour_df[~tour_df["sap"].isin(SELECTED_SAPS) & tour_df["sap"].ne("")]
    if unknown.empty:
        return pd.DataFrame(columns=empty_cols)

    agg = unknown.groupby("sap", as_index=False).agg(
        blaetter=("blatt", lambda x: sorted(set(x))),
        gesamt=("blatt", "size"),
    )
    agg["SAP Nummer"] = agg["sap"]
    agg["Blätter Tourenplanung"] = agg["blaetter"].map(", ".join)
    agg["Gesamt-Vorkommen"] = agg["gesamt"]
    agg = agg.sort_values("SAP Nummer").reset_index(drop=True)
    return agg[empty_cols]


def _export_columns_missing() -> List[str]:
    return [
        "Standort",
        "SAP Nummer",
        "Fehlende Liefertage",
        "Anzahl fehlender Tage",
        "Blätter Tourenplanung",
        "Liefertage in SAP",
        "Hinweis",
    ]


def _export_columns_missing_tour() -> List[str]:
    return [
        "Standort",
        "SAP Nummer",
        "Fehlende Liefertage",
        "Anzahl fehlender Tage",
        "Liefertage in Tourenplanung",
        "Hinweis",
    ]


def _empty_result_df() -> pd.DataFrame:
    return pd.DataFrame(columns=_export_columns_missing())


def interleave_blank_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Fügt nach jeder Datenzeile außer der letzten eine komplett leere Zeile ein."""
    if df.empty or len(df) <= 1:
        return df.reset_index(drop=True)
    blank = pd.DataFrame([[pd.NA] * len(df.columns)], columns=df.columns)
    pieces: List[pd.DataFrame] = []
    for i in range(len(df)):
        pieces.append(df.iloc[[i]])
        if i < len(df) - 1:
            pieces.append(blank)
    return pd.concat(pieces, ignore_index=True)


def build_excel(
    missing_sap: pd.DataFrame,
    missing_tour: pd.DataFrame | None,
    unknown_saps: pd.DataFrame | None,
) -> bytes:
    """Schreibt eine formatierte Excel-Datei mit bis zu drei Blättern.
    Zwischen jedem Kunden wird eine Leerzeile eingefügt."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        missing_sap_out = interleave_blank_rows(missing_sap)
        missing_sap_out.to_excel(writer, index=False, sheet_name="Fehlt in SAP", na_rep="")
        _format_sheet(writer, "Fehlt in SAP", missing_sap_out)

        if missing_tour is not None:
            missing_tour_out = interleave_blank_rows(missing_tour)
            missing_tour_out.to_excel(writer, index=False, sheet_name="Fehlt in Tourenplanung", na_rep="")
            _format_sheet(writer, "Fehlt in Tourenplanung", missing_tour_out)

        if unknown_saps is not None and not unknown_saps.empty:
            unknown_out = interleave_blank_rows(unknown_saps)
            unknown_out.to_excel(writer, index=False, sheet_name="Unbekannte SAP-Nummern", na_rep="")
            _format_sheet(writer, "Unbekannte SAP-Nummern", unknown_out)

    return output.getvalue()


def _format_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    if df is None:
        return
    ws = writer.sheets[sheet_name]

    # Header formatieren
    header_fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_align = Alignment(horizontal="left", vertical="center")
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Spaltenbreiten grob an Inhaltslänge
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col_name))] +
            [len(str(v)) for v in df[col_name].astype(str).head(200).tolist()]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)

    # Kopfzeile einfrieren + Autofilter
    ws.freeze_panes = "A2"
    if not df.empty:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"


def build_group_overview() -> str:
    parts: List[str] = []
    for location_name, sap_list in CUSTOMER_GROUPS.items():
        parts.append(f"{location_name} ({len(sap_list)} Kunden)")
        parts.append("\n".join(sap_list))
        parts.append("")
    return "\n".join(parts).strip()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Tourenplanung gegen SAP", layout="wide")

st.title("Tourenplanung gegen SAP")
st.write(
    "Vergleicht die Liefertage in der Tourenplanung gegen die in SAP hinterlegten Liefertage "
    "für die drei Standorte Malchow, Neumünster und Zarrentin."
)

# Datenqualitäts-Warnung bei Duplikaten in der Konfiguration
duplicates = find_duplicate_saps()
if duplicates:
    with st.expander(f"⚠️ {len(duplicates)} doppelte SAP-Nummer(n) in der Kundensortierung", expanded=False):
        for sap, locations in duplicates:
            st.write(f"- **{sap}**: {', '.join(locations)}")

st.info(
    "Richtung des Vergleichs:\n"
    "- SAP = Datei mit SAP Nummer in A und Liefertag in G\n"
    "- Tourenplanung = Datei mit Spalte B sowie Montag bis Samstag in G bis L\n"
    "- Standard-Ausgabe = nur Tage, die in der Tourenplanung stehen, aber in SAP fehlen\n"
    "- Sortierung = zuerst Malchow, dann Neumünster, dann Zarrentin"
)

col1, col2, col3 = st.columns(3)
col1.metric("Malchow", len(CUSTOMER_GROUPS["Malchow"]))
col2.metric("Neumünster", len(CUSTOMER_GROUPS["Neumünster"]))
col3.metric("Zarrentin", len(CUSTOMER_GROUPS["Zarrentin"]))

with st.expander("Hinterlegte Kundensortierung", expanded=False):
    st.text_area(
        "Die Kunden werden mit dieser Reihenfolge ausgewertet und in der Excel genauso sortiert.",
        value=build_group_overview(),
        height=420,
        disabled=True,
    )

sap_datei = st.file_uploader(
    "SAP hochladen – erstes Blatt, Spalte A = SAP Nummer, Spalte G = Liefertag 1 bis 6",
    type=["xlsx", "xlsm", "xls"],
    key="sap_datei",
)

tourenplanung_datei = st.file_uploader(
    "Tourenplanung hochladen – erste 4 Blätter, Spalte B = SAP Nummer, Spalte G bis L = Montag bis Samstag",
    type=["xlsx", "xlsm", "xls"],
    key="tourenplanung_datei",
)

with st.expander("Optionen", expanded=False):
    include_reverse = st.checkbox(
        "Zusätzlich prüfen: Tage, die in SAP stehen, aber in der Tourenplanung fehlen (eigenes Blatt)",
        value=False,
    )
    include_unknown = st.checkbox(
        "Unbekannte SAP-Nummern aus der Tourenplanung mit ausgeben (eigenes Blatt)",
        value=True,
    )

run = st.button("Excel erzeugen", type="primary")

if run:
    if not sap_datei or not tourenplanung_datei:
        st.error("Bitte beide Excel-Dateien hochladen.")
        st.stop()

    try:
        days_by_sap, sap_sheet, sap_rows = read_sap_file(sap_datei)
        tour_df, tour_sheets = read_tourenplanung(tourenplanung_datei)

        missing_sap = build_missing_in_sap(tour_df, days_by_sap)
        missing_tour = build_missing_in_tour(tour_df, days_by_sap) if include_reverse else None
        unknown_saps = build_unknown_saps(tour_df) if include_unknown else None

        excel_bytes = build_excel(missing_sap, missing_tour, unknown_saps)

        # State halten, damit Download-Button kein Re-Run auslöst
        st.session_state["result"] = {
            "missing_sap": missing_sap,
            "missing_tour": missing_tour,
            "unknown_saps": unknown_saps,
            "excel_bytes": excel_bytes,
            "sap_sheet": sap_sheet,
            "sap_rows": sap_rows,
            "tour_sheets": tour_sheets,
        }
    except Exception as exc:
        st.error(f"Fehler beim Verarbeiten der Dateien: {exc}")
        st.session_state.pop("result", None)

# Ergebnisanzeige
result = st.session_state.get("result")
if result:
    missing_sap = result["missing_sap"]
    missing_tour = result["missing_tour"]
    unknown_saps = result["unknown_saps"]

    st.success(
        f"Fertig. {len(missing_sap)} Kunden mit Tagen in der Tourenplanung, "
        f"die in SAP als Liefertag fehlen."
    )

    st.caption(
        f"SAP: Blatt = {result['sap_sheet']}, {result['sap_rows']} Liefertage übernommen | "
        f"Tourenplanung: geprüfte Blätter = {', '.join(result['tour_sheets'])}"
    )

    if not missing_sap.empty:
        standort_zaehlung = (
            missing_sap.groupby("Standort", as_index=False)
            .agg(
                **{
                    "Betroffene Kunden": ("SAP Nummer", "count"),
                    "Fehlende Tage gesamt": ("Anzahl fehlender Tage", "sum"),
                }
            )
        )
        st.dataframe(standort_zaehlung, use_container_width=True, hide_index=True)

        with st.expander("Vorschau: Fehlt in SAP", expanded=False):
            standorte = ["Alle"] + sorted(missing_sap["Standort"].unique().tolist())
            filter_wahl = st.selectbox("Standort filtern", standorte, key="filter_missing_sap")
            vorschau = missing_sap if filter_wahl == "Alle" else missing_sap[missing_sap["Standort"] == filter_wahl]
            st.dataframe(vorschau, use_container_width=True, hide_index=True)

    if missing_tour is not None and not missing_tour.empty:
        with st.expander(f"Vorschau: Fehlt in Tourenplanung ({len(missing_tour)} Kunden)", expanded=False):
            st.dataframe(missing_tour, use_container_width=True, hide_index=True)

    if unknown_saps is not None and not unknown_saps.empty:
        with st.expander(f"Vorschau: Unbekannte SAP-Nummern ({len(unknown_saps)} Kunden)", expanded=False):
            st.dataframe(unknown_saps, use_container_width=True, hide_index=True)

    st.download_button(
        label="Excel herunterladen",
        data=result["excel_bytes"],
        file_name="tourenplanung_tage_fehlen_in_sap_sortiert.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
